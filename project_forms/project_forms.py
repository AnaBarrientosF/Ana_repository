import streamlit as st
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
import time

# 🔹 Ruta local en OneDrive (lectura y escritura)
LOCAL_PATH = Path(r"C:\Users\Ana Barrientos\OneDrive - Grupoesis\Dashboard proyectos\Proyectos nacionales\Base de datos Proyectos.xlsx")

def cargar_df_prueba(path: Path, reintentos: int = 10, espera_seg: float = 0.5):
    """Leer la hoja 'Prueba' reintentando si el archivo está bloqueado."""
    for i in range(reintentos):
        try:
            return pd.read_excel(path, sheet_name="Prueba", engine="openpyxl")
        except PermissionError:
            if i == reintentos - 1:
                st.error("❌ El archivo está en uso. Cierra el Excel y espera a que OneDrive termine de sincronizar.")
                st.stop()
            time.sleep(espera_seg)

# ⭐ NUEVO: usar session_state para que df no se pierda al recargar
if "df_prueba" not in st.session_state:
    st.session_state["df_prueba"] = cargar_df_prueba(LOCAL_PATH)

# 👉 Usamos el df desde session_state
df = st.session_state["df_prueba"]

st.set_page_config(page_title="Formulario de Proyectos", layout="centered")
st.title("Formulario de Proyectos Nacionales")

ingenieros = ["Justin Aguirre", "Erick Villalobos", "Guillermo Ordoñez", "Dylan López"]

with st.form("form_registro"):
    cliente = st.text_input("Cliente")
    nombre_proyecto = st.text_input("Nombre del Proyecto")
    ingeniero = st.selectbox("Ingeniero de implementación", ingenieros)
    enviado = st.form_submit_button("Guardar registro")

if enviado:
    if not cliente or not nombre_proyecto:
        st.error("⚠️ Los campos Cliente y Nombre del Proyecto son obligatorios.")
    else:
        # Asegurar la columna NO
        if "NO" not in df.columns:
            df.insert(0, "NO", list(range(1, len(df) + 1)))

        next_id = int(df["NO"].max()) + 1 if not df.empty else 1

        # Actualizar df en memoria
        nueva_fila_df = {
            "NO": next_id,
            "CLIENTE": cliente,
            "PROYECTO": nombre_proyecto,
            "INGENIERO DE IMPLEMENTACION": ingeniero
        }
        df = pd.concat([df, pd.DataFrame([nueva_fila_df])], ignore_index=True)

        # ⭐ NUEVO: actualizar lo que queda guardado en session_state
        st.session_state["df_prueba"] = df

        # Intentar abrir el archivo para escritura (con reintentos)
        for i in range(10):
            try:
                wb = load_workbook(LOCAL_PATH)
                break
            except PermissionError:
                if i == 9:
                    st.error("❌ No se pudo escribir en el archivo porque está en uso. Cierra el Excel y espera a que OneDrive termine de sincronizar.")
                    st.stop()
                time.sleep(0.5)

        if "Prueba" not in wb.sheetnames:
            st.error("❌ La hoja 'Prueba' no existe en el archivo.")
        else:
            ws = wb["Prueba"]

            nueva_fila_excel = [
                next_id,
                cliente,
                nombre_proyecto,
                ingeniero
            ]

            ws.append(nueva_fila_excel)
            wb.save(LOCAL_PATH)

            st.success(f"✅ Registro agregado a la hoja 'Prueba' (ID: {next_id}).")

# Mostrar tabla SIEMPRE basada en el df actual (que incluye lo que acabas de agregar)
st.subheader("📄 Registros existentes (Hoja: Prueba)")
st.dataframe(df, use_container_width=True)
